#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
微信图片解密密钥提取工具
从微信缓存的模板文件中提取XOR密钥和AES密钥

参考: https://github.com/recarto404/WxDatDecrypt
"""

import os
import struct
from pathlib import Path
from collections import Counter
from typing import Optional, Tuple, List
import openpyxl

try:
    from Crypto.Cipher import AES
    from Crypto.Util.Padding import unpad
    HAS_CRYPTO = True
except ImportError:
    HAS_CRYPTO = False

try:
    import pymem
    import pymem.process
    HAS_PYMEM = True
except ImportError:
    HAS_PYMEM = False


def find_xor_key(dat_directory: str) -> Optional[int]:
    """
    从微信缓存目录的模板文件(*_t.dat)中提取XOR密钥

    原理:
    微信的缩略图文件(*_t.dat)是标准JPEG格式加密而来
    JPEG文件末尾固定为 0xFF 0xD9
    通过分析多个模板文件的最后2字节,找到最常见的加密字节对
    验证: x ^ 0xFF == y ^ 0xD9
    密钥为: x ^ 0xFF 或 y ^ 0xD9

    参数:
        dat_directory: 微信图片缓存目录(包含.dat文件)

    返回:
        XOR密钥(0-255) 或 None
    """
    dat_dir = Path(dat_directory)
    if not dat_dir.exists():
        print(f"[ERROR] 目录不存在: {dat_directory}")
        return None

    # 查找所有缩略图文件 (*_t.dat)
    template_files = list(dat_dir.glob('*_t.dat'))

    if len(template_files) < 5:
        print(f"[WARNING] 缩略图文件太少 ({len(template_files)}个), 至少需要5个")
        print("[INFO] 尝试查找普通.dat文件...")
        template_files = list(dat_dir.glob('*.dat'))[:16]

    if len(template_files) == 0:
        print(f"[ERROR] 未找到.dat文件: {dat_directory}")
        return None

    print(f"[INFO] 找到 {len(template_files)} 个模板文件")

    # 读取每个文件的最后2字节
    last_bytes = []
    for dat_file in template_files[:16]:  # 最多检查16个文件
        try:
            with open(dat_file, 'rb') as f:
                f.seek(-2, 2)  # 从文件末尾往前2字节
                last_two = f.read(2)
                if len(last_two) == 2:
                    last_bytes.append(tuple(last_two))
        except Exception as e:
            continue

    if len(last_bytes) < 3:
        print(f"[ERROR] 可读取的文件太少")
        return None

    # 统计最常见的字节对
    counter = Counter(last_bytes)
    most_common = counter.most_common(1)[0]
    byte_pair = most_common[0]
    count = most_common[1]

    x, y = byte_pair
    print(f"[INFO] 最常见的末尾字节对: 0x{x:02X} 0x{y:02X} (出现{count}次)")

    # 验证: JPEG末尾应该是 0xFF 0xD9
    # 加密后: x = 0xFF ^ key, y = 0xD9 ^ key
    # 因此: x ^ 0xFF == y ^ 0xD9 == key

    key1 = x ^ 0xFF
    key2 = y ^ 0xD9

    if key1 == key2:
        print(f"[OK] 找到XOR密钥: {key1} (0x{key1:02X})")
        return key1
    else:
        print(f"[WARNING] 字节对验证失败: key1={key1}, key2={key2}")
        print(f"[INFO] 使用key1: {key1}")
        return key1


def find_v4_template_file(dat_directory: str) -> Optional[str]:
    """
    查找v4格式的模板文件（用于AES密钥验证）

    v4格式文件特征：
    - 文件头: b"\x07\x08V2\x08\x07"

    参数:
        dat_directory: 包含.dat文件的目录

    返回:
        v4模板文件路径 或 None
    """
    dat_dir = Path(dat_directory)
    if not dat_dir.exists():
        return None

    # 查找所有.dat文件
    for dat_file in dat_dir.glob('*.dat'):
        try:
            with open(dat_file, 'rb') as f:
                header = f.read(6)
                if header == b"\x07\x08V2\x08\x07":
                    return str(dat_file)
        except:
            continue

    return None


def extract_aes_key_from_process(process_name: str = "WeChat.exe") -> Optional[List[str]]:
    """
    从微信进程内存中提取AES密钥

    原理:
    1. 附加到WeChat.exe进程
    2. 扫描进程内存，查找16字节密钥候选
    3. 使用简单模式匹配（连续可打印字符或特定模式）

    参数:
        process_name: 进程名称

    返回:
        密钥候选列表(十六进制字符串) 或 None
    """
    if not HAS_PYMEM:
        print("[ERROR] 需要安装pymem库")
        print("[INFO] 安装命令: pip install pymem")
        return None

    try:
        # 附加到进程
        pm = pymem.Pymem(process_name)
        print(f"[INFO] 已附加到进程: {process_name} (PID: {pm.process_id})")

        # 获取进程所有模块
        modules = list(pm.list_modules())
        print(f"[INFO] 找到 {len(modules)} 个模块")

        candidates = []

        # 扫描每个模块的内存
        for module in modules:
            try:
                # 只扫描主模块和相关DLL
                if not any(name in module.name.lower() for name in ['wechat', 'core', 'util']):
                    continue

                print(f"[INFO] 扫描模块: {module.name}")

                # 读取模块内存
                base_address = module.lpBaseOfDll
                module_size = module.SizeOfImage

                # 分块读取（避免一次性读取过大内存）
                chunk_size = 1024 * 1024  # 1MB

                for offset in range(0, module_size, chunk_size):
                    try:
                        read_size = min(chunk_size, module_size - offset)
                        data = pm.read_bytes(base_address + offset, read_size)

                        # 搜索16字节的密钥模式
                        for i in range(len(data) - 16):
                            key_candidate = data[i:i+16]

                            # 简单启发式：检查是否有一定的随机性
                            # 避免全0、全FF等无效密钥
                            if (key_candidate == b'\x00' * 16 or
                                key_candidate == b'\xff' * 16):
                                continue

                            # 转换为十六进制
                            key_hex = key_candidate.hex()

                            # 避免重复
                            if key_hex not in candidates:
                                candidates.append(key_hex)

                                # 限制候选数量（避免过多）
                                if len(candidates) >= 1000:
                                    break

                        if len(candidates) >= 1000:
                            break

                    except Exception as e:
                        continue

                if len(candidates) >= 1000:
                    break

            except Exception as e:
                continue

        print(f"[INFO] 找到 {len(candidates)} 个密钥候选")
        return candidates if len(candidates) > 0 else None

    except Exception as e:
        print(f"[ERROR] 进程访问失败: {e}")
        return None


def verify_aes_key(template_file: str, aes_key: str, xor_key: int) -> bool:
    """
    验证AES密钥是否正确

    参数:
        template_file: v4格式的模板文件路径
        aes_key: AES密钥(32字节十六进制字符串)
        xor_key: XOR密钥

    返回:
        True: 密钥正确
        False: 密钥错误
    """
    if not HAS_CRYPTO:
        return False

    try:
        with open(template_file, 'rb') as f:
            data = f.read()

        # 解析文件头
        if data[:6] != b"\x07\x08V2\x08\x07":
            return False

        # 读取大小信息
        aes_size = struct.unpack('<I', data[6:10])[0]
        xor_size = struct.unpack('<I', data[10:14])[0]

        # 转换密钥
        try:
            aes_key_bytes = bytes.fromhex(aes_key)
        except:
            return False

        if len(aes_key_bytes) != 16:
            return False

        # 读取加密数据
        encrypted_data = data[15:]
        aes_padded_size = (aes_size + 15) // 16 * 16
        aes_data = encrypted_data[:aes_padded_size]

        # AES解密
        cipher = AES.new(aes_key_bytes, AES.MODE_ECB)
        decrypted = cipher.decrypt(aes_data)

        # 去除填充
        decrypted = decrypted[:aes_size]

        # 验证JPEG魔术字节
        if decrypted[:3] == b'\xFF\xD8\xFF':
            return True

        return False

    except Exception:
        return False


def find_aes_key(dat_directory: str, xor_key: int, process_name: str = "WeChat.exe") -> Optional[str]:
    """
    自动查找AES密钥

    流程:
    1. 查找v4格式的模板文件
    2. 从进程内存提取密钥候选
    3. 逐个验证密钥

    参数:
        dat_directory: 包含.dat文件的目录
        xor_key: XOR密钥(用于验证)
        process_name: 微信进程名称

    返回:
        AES密钥(十六进制字符串) 或 None
    """
    print("\n[INFO] 开始查找AES密钥...")

    # 1. 查找v4模板文件
    template_file = find_v4_template_file(dat_directory)
    if not template_file:
        print("[ERROR] 未找到v4格式的模板文件")
        print("[INFO] 可能原因：")
        print("  - 该目录下没有v4格式的加密文件")
        print("  - 微信版本较旧，使用的是v3格式（仅需XOR密钥）")
        return None

    print(f"[OK] 找到v4模板文件: {Path(template_file).name}")

    # 2. 从进程提取密钥候选
    print("\n[INFO] 从微信进程提取密钥...")
    print("[WARNING] 此操作需要管理员权限")

    candidates = extract_aes_key_from_process(process_name)
    if not candidates:
        print("[ERROR] 未能从进程提取密钥")
        return None

    print(f"[INFO] 提取到 {len(candidates)} 个密钥候选")

    # 3. 验证密钥
    print("\n[INFO] 验证密钥...")
    for i, key in enumerate(candidates):
        if (i + 1) % 100 == 0:
            print(f"[INFO] 已验证 {i+1}/{len(candidates)} 个候选...")

        if verify_aes_key(template_file, key, xor_key):
            print(f"\n[OK] 找到有效的AES密钥: {key}")
            return key

    print("\n[ERROR] 未找到有效的AES密钥")
    return None


def verify_xor_key(dat_file: str, xor_key: int) -> bool:
    """
    验证XOR密钥是否正确

    参数:
        dat_file: .dat文件路径
        xor_key: XOR密钥

    返回:
        True: 解密后是有效的JPEG文件
        False: 解密失败
    """
    try:
        with open(dat_file, 'rb') as f:
            encrypted = f.read(10)  # 只读前10字节验证

        decrypted = bytes([b ^ xor_key for b in encrypted])

        # 检查JPEG魔术字节: FF D8 FF
        if decrypted[:3] == b'\xFF\xD8\xFF':
            return True

        return False

    except Exception:
        return False


def save_keys_to_config(config_path: str, xor_key: Optional[int] = None,
                        aes_key: Optional[str] = None) -> bool:
    """
    保存密钥到config.xlsx

    参数:
        config_path: config.xlsx路径
        xor_key: XOR密钥
        aes_key: AES密钥(v4格式)

    返回:
        True: 成功, False: 失败
    """
    try:
        config_path = Path(config_path)

        # 如果配置文件存在,读取现有配置
        if config_path.exists():
            wb = openpyxl.load_workbook(config_path)
            ws = wb['config'] if 'config' in wb.sheetnames else wb.active
        else:
            # 创建新配置文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'config'
            # 添加表头
            ws['A1'] = 'key'
            ws['B1'] = 'value'

        # 查找或添加密钥行
        xor_row = None
        aes_row = None

        for row in range(2, ws.max_row + 1):
            key_name = ws[f'A{row}'].value
            if key_name == 'image_xor_key':
                xor_row = row
            elif key_name == 'image_aes_key':
                aes_row = row

        # 保存XOR密钥
        if xor_key is not None:
            if xor_row is None:
                xor_row = ws.max_row + 1
                ws[f'A{xor_row}'] = 'image_xor_key'
            ws[f'B{xor_row}'] = str(xor_key)

        # 保存AES密钥
        if aes_key is not None:
            if aes_row is None:
                aes_row = ws.max_row + 1
                ws[f'A{aes_row}'] = 'image_aes_key'
            ws[f'B{aes_row}'] = aes_key

        # 保存文件
        wb.save(config_path)
        print(f"[OK] 密钥已保存到: {config_path}")
        return True

    except Exception as e:
        print(f"[ERROR] 保存配置失败: {e}")
        return False


def load_keys_from_config(config_path: str) -> Tuple[Optional[int], Optional[str]]:
    """
    从config.xlsx读取密钥

    参数:
        config_path: config.xlsx路径

    返回:
        (xor_key, aes_key) 或 (None, None)
    """
    try:
        config_path = Path(config_path)
        if not config_path.exists():
            return None, None

        wb = openpyxl.load_workbook(config_path)
        ws = wb['config'] if 'config' in wb.sheetnames else wb.active

        xor_key = None
        aes_key = None

        for row in range(2, ws.max_row + 1):
            key_name = ws[f'A{row}'].value
            key_value = ws[f'B{row}'].value

            if key_name == 'image_xor_key' and key_value:
                try:
                    xor_key = int(key_value)
                except:
                    pass
            elif key_name == 'image_aes_key' and key_value:
                aes_key = str(key_value)

        return xor_key, aes_key

    except Exception as e:
        print(f"[ERROR] 读取配置失败: {e}")
        return None, None


def auto_extract_keys(wechat_files_dir: str, config_path: str = 'config.xlsx') -> Tuple[Optional[int], Optional[str]]:
    """
    自动从微信文件目录提取密钥并保存到配置

    参数:
        wechat_files_dir: 微信文件根目录 (如: C:/Users/xxx/xwechat_files/账号)
        config_path: 配置文件路径

    返回:
        (xor_key, aes_key) 或 (None, None)
    """
    print("="*70)
    print("微信图片解密密钥提取")
    print("="*70)
    print()

    wechat_dir = Path(wechat_files_dir)
    if not wechat_dir.exists():
        print(f"[ERROR] 微信文件目录不存在: {wechat_files_dir}")
        return None, None

    # 查找图片缓存目录
    # 常见位置: msg/attach/xxxxxx/YYYY-MM/Img
    possible_img_dirs = []

    # 方法1: 搜索msg/attach下的Img目录
    attach_dir = wechat_dir / 'msg' / 'attach'
    if attach_dir.exists():
        for img_dir in attach_dir.rglob('Img'):
            if img_dir.is_dir():
                possible_img_dirs.append(img_dir)

    # 方法2: 搜索image2目录(旧版本)
    image2_dir = wechat_dir / 'image2'
    if image2_dir.exists():
        possible_img_dirs.append(image2_dir)

    if len(possible_img_dirs) == 0:
        print(f"[ERROR] 未找到图片缓存目录")
        print("[INFO] 请手动指定包含.dat文件的目录")
        return None, None

    print(f"[INFO] 找到 {len(possible_img_dirs)} 个候选目录")

    # 尝试从每个目录提取密钥
    xor_key = None
    for img_dir in possible_img_dirs:
        print(f"\n[INFO] 检查目录: {img_dir}")
        xor_key = find_xor_key(str(img_dir))
        if xor_key is not None:
            # 验证密钥
            dat_files = list(img_dir.glob('*.dat'))
            if len(dat_files) > 0:
                if verify_xor_key(str(dat_files[0]), xor_key):
                    print(f"[OK] 密钥验证成功")
                    break
                else:
                    print(f"[WARNING] 密钥验证失败，继续查找...")
                    xor_key = None

    if xor_key is None:
        print("\n[ERROR] 未能提取XOR密钥")
        return None, None

    # AES密钥提取(v4格式,需要进程内存访问)
    aes_key = None

    # 检查是否有v4格式文件
    has_v4 = False
    for img_dir in possible_img_dirs:
        v4_file = find_v4_template_file(str(img_dir))
        if v4_file:
            has_v4 = True
            print(f"\n[INFO] 检测到v4格式文件，尝试提取AES密钥...")

            if not HAS_PYMEM:
                print("[WARNING] 未安装pymem库，无法自动提取AES密钥")
                print("[INFO] 安装命令: pip install pymem")
                print("[INFO] 或手动提供AES密钥")
            elif not HAS_CRYPTO:
                print("[WARNING] 未安装pycryptodome库，无法验证AES密钥")
                print("[INFO] 安装命令: pip install pycryptodome")
            else:
                # 尝试提取AES密钥
                try:
                    aes_key = find_aes_key(str(img_dir), xor_key)
                    if aes_key:
                        print(f"[OK] 成功提取AES密钥")
                    else:
                        print("[WARNING] 未能自动提取AES密钥")
                        print("[INFO] 请确保：")
                        print("  1. 微信正在运行")
                        print("  2. 以管理员权限运行此脚本")
                        print("  3. 或手动提供AES密钥")
                except Exception as e:
                    print(f"[WARNING] AES密钥提取失败: {e}")
            break

    if not has_v4:
        print("\n[INFO] 未检测到v4格式文件，仅需XOR密钥即可")

    # 保存到配置
    print()
    save_keys_to_config(config_path, xor_key, aes_key)

    print()
    print("="*70)
    print("密钥提取完成")
    print("="*70)
    print(f"XOR密钥: {xor_key} (0x{xor_key:02X})")
    if aes_key:
        print(f"AES密钥: {aes_key}")
    print()

    return xor_key, aes_key


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("用法:")
        print("  python wechat_key_extractor.py <微信文件目录> [config.xlsx路径]")
        print("\n示例:")
        print("  python wechat_key_extractor.py C:/Users/用户名/xwechat_files/账号")
        sys.exit(1)

    wechat_dir = sys.argv[1]
    config_file = sys.argv[2] if len(sys.argv) > 2 else 'config.xlsx'

    auto_extract_keys(wechat_dir, config_file)
